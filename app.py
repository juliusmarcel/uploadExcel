from flask import Flask, request, jsonify, render_template
import pandas as pd
import pyodbc
from openpyxl import load_workbook

app = Flask(__name__)

# Konfigurasi Koneksi ke SQL Server
DB_CONN = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=LAPTOP-FDEOFA5T;"
    "DATABASE=UploadExcelDB;"
    "Trusted_Connection=yes;"
)

# Fungsi untuk memetakan tipe data Pandas ke SQL Server
def map_dtype_to_sql(dtype):
    if pd.api.types.is_integer_dtype(dtype):
        return "INT"
    elif pd.api.types.is_float_dtype(dtype):
        return "FLOAT"
    elif pd.api.types.is_datetime64_any_dtype(dtype):
        return "DATETIME"
    else:
        return "NVARCHAR(255)"  # Default untuk teks

# Fungsi untuk membuat tabel secara otomatis
def create_table_if_not_exists(table_name, df):
    conn = pyodbc.connect(DB_CONN)
    cursor = conn.cursor()

    # Cek apakah tabel sudah ada
    check_query = f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table_name}'"
    cursor.execute(check_query)
    if cursor.fetchone()[0] > 0:
        print(f"Tabel '{table_name}' sudah ada.")
        cursor.close()
        conn.close()
        return

    # Buat tabel jika belum ada
    columns = ", ".join([f"[{col}] {map_dtype_to_sql(df[col].dtype)}" for col in df.columns])
    create_query = f"CREATE TABLE [{table_name}] (ID INT IDENTITY(1,1) PRIMARY KEY, {columns});"

    try:
        cursor.execute(create_query)
        conn.commit()
        print(f"Tabel '{table_name}' berhasil dibuat.")
    except Exception as e:
        print(f"Error membuat tabel '{table_name}': {e}")

    cursor.close()
    conn.close()

# Fungsi untuk menyimpan data ke SQL Server
def save_to_db(df, table_name):
    conn = pyodbc.connect(DB_CONN)
    cursor = conn.cursor()

    for _, row in df.iterrows():
        placeholders = ", ".join(["?" for _ in df.columns])
        columns = ", ".join([f"[{col}]" for col in df.columns])
        query = f"INSERT INTO [{table_name}] ({columns}) VALUES ({placeholders})"

        try:
            cursor.execute(query, tuple(row))
        except Exception as e:
            print(f"Error insert ke '{table_name}': {e}")

    conn.commit()
    cursor.close()
    conn.close()

# Fungsi untuk membaca tabel dari file Excel (menggunakan openpyxl)
def read_tables_from_excel(file):
    wb = load_workbook(file, data_only=True)
    tables = []

    for sheet in wb.sheetnames:
        sheet_obj = wb[sheet]
        for table in sheet_obj.tables.values():
            table_range = sheet_obj[table.ref]
            # Mengambil data tabel sebagai DataFrame
            data = []
            for row in table_range:
                data.append([cell.value for cell in row])
            tables.append((sheet, table.name, pd.DataFrame(data[1:], columns=data[0])))

    return tables

# API untuk upload file Excel
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    
    # Membaca tabel-tabel dari file Excel
    tables = read_tables_from_excel(file)

    result = {}

    for sheet_name, table_name, df in tables:
        # Buat tabel jika belum ada
        create_table_if_not_exists(table_name, df)

        # Simpan data ke SQL Server
        save_to_db(df, table_name)

        result[table_name] = f"{len(df)} rows uploaded"

    return jsonify({"message": "Upload successful", "details": result}), 200

# Route untuk menyajikan index.html
@app.route('/')
def index():
    return render_template('index.html')

# Jalankan Flask
if __name__ == '__main__':
    app.run(debug=True)
    